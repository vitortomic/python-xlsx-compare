import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

red_fill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')

first_dir = os.listdir('./first')
second_dir = os.listdir('./second')

if len(first_dir) < 1 or len(second_dir) < 1:
    raise Exception('Missing files!')

first_file_name = first_dir[0]
second_file_name = second_dir[0]

print(first_file_name)
print(second_file_name)

wb1 = load_workbook(filename = f'./first/{first_file_name}')
wb2 = load_workbook(filename = f'./second/{second_file_name}')

sheet_name_first = wb1.sheetnames[0]
sheet_name_second = wb2.sheetnames[0]

if sheet_name_first != sheet_name_second:
    raise Exception('Reports are not of the same type!')

ws1 = wb1[sheet_name_first]
ws2 = wb2[sheet_name_second]

if ws1.max_column != ws2.max_column:
    raise Exception('Number of columns does not match!')

if ws1.max_row != ws2.max_row:
    print('Number of rows does not match!')

header_row = 6
max_row = min(ws1.max_row, ws2.max_row)
number_of_differences = 0


for i in range (header_row + 1, ws1.max_row - 1):
    for j in range (1, ws1.max_column):
        if ws1.cell(i, j).value != ws2.cell(i, j).value:
            number_of_differences += 1
            print(f'Difference found for column "{ws1.cell(header_row, j).value}"')
            print(f'Value for first file: {ws1.cell(i, j).value}')
            print(f'Value for second file: {ws2.cell(i, j).value}')
            ws1.cell(i, j).fill = red_fill
            ws2.cell(i, j).fill = red_fill

if number_of_differences > 0:
    print(f'Marked {number_of_differences} differences in spreadsheets.')
    wb1.save(filename = f'./first/{first_file_name}')
    wb2.save(filename = f'./second/{second_file_name}')
else:
    print('No differences between spreadsheets.')
